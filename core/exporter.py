"""
Excel导出模块

负责将计算结果导出为格式化的Excel文件
"""

from datetime import datetime
from pathlib import Path
from typing import Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from core.models import Table1Result, Table2Result
from core.constants import (
    TRANSACTION_DECIMALS,
    EXPENSE_DECIMALS,
    PERCENTAGE_DECIMALS,
    TABLE1_COLUMNS
)


class ExporterError(Exception):
    """导出错误"""
    pass


def export_to_excel(
    table1: Table1Result,
    table2: Table2Result,
    output_path: Optional[str] = None
) -> str:
    """
    导出计算结果到Excel文件

    Args:
        table1: 渠道维度预算表结果
        table2: 客群维度汇总表结果
        output_path: 输出文件路径 (可选, 默认自动生成)

    Returns:
        str: 输出文件的绝对路径

    Raises:
        ExporterError: 导出失败时抛出
    """
    try:
        is_filelike = hasattr(output_path, 'write')

        if not is_filelike:
            # 生成默认文件名
            if output_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"预算推算结果_{timestamp}.xlsx"
                output_path = str(Path.cwd() / "export" / filename)

            # 确保输出目录存在
            output_dir = Path(output_path).parent
            output_dir.mkdir(parents=True, exist_ok=True)

        # 转换为DataFrame
        df1 = table1.to_dataframe()
        df2 = table2.to_dataframe()

        # 应用格式化
        df1_formatted = _format_dataframe(df1, sheet_type="table1")
        df2_formatted = _format_dataframe(df2, sheet_type="table2")

        # 写入Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df1_formatted.to_excel(writer, sheet_name='渠道维度预算表', index=False)
            df2_formatted.to_excel(writer, sheet_name='客群维度汇总表', index=False)

        # 应用样式 (only for file paths, not file-like objects)
        if not is_filelike:
            _apply_excel_styles(output_path)
            return str(Path(output_path).resolve())

        return output_path

    except Exception as e:
        raise ExporterError(f"导出Excel失败: {e}")


def _format_dataframe(df: pd.DataFrame, sheet_type: str) -> pd.DataFrame:
    """
    格式化DataFrame的数值

    Args:
        df: 待格式化的DataFrame
        sheet_type: 表类型 ("table1" 或 "table2")

    Returns:
        pd.DataFrame: 格式化后的DataFrame
    """
    df = df.copy()

    if sheet_type == "table1":
        # 花费: 万元/千万元, 取整（无小数）—— 匹配 Table1 to_dataframe() 输出的列名
        expense_col = '花费(千万元)' if '花费(千万元)' in df.columns else ('花费(万元)' if '花费(万元)' in df.columns else '花费')
        if expense_col in df.columns:
            df[expense_col] = df[expense_col].apply(
                lambda x: f"{int(round(x)):,}" if pd.notna(x) else ""
            )

        # 交易额: 千万元, 2位小数
        transaction_cols = ['T0交易额(千万元)', '当月首登M0交易额(千万元)', 'T0交易额', '当月首登M0交易额']
        for col in transaction_cols:
            if col in df.columns:
                df[col] = df[col].apply(
                    lambda x: f"{x:.{TRANSACTION_DECIMALS}f}" if pd.notna(x) else ""
                )

        # 百分比: XX.X%, 1位小数
        percentage_cols = ['花费结构', '申完结构', '1-3 T0过件率', '1-8 T0CPS']
        for col in percentage_cols:
            if col in df.columns:
                df[col] = df[col].apply(
                    lambda x: f"{x*100:.{PERCENTAGE_DECIMALS}f}%" if pd.notna(x) and isinstance(x, float) and x <= 1 else (
                        f"{x:.{PERCENTAGE_DECIMALS}f}%" if pd.notna(x) else ""
                    )
                )

        # 其他数值列保留2位小数
        numeric_cols = ['非年龄拒绝申完量', '1-3 T0授信量', 'T0申完量']
        for col in numeric_cols:
            if col in df.columns:
                df[col] = df[col].apply(
                    lambda x: f"{x:,.0f}" if pd.notna(x) else ""
                )

    elif sheet_type == "table2":
        # 第二张表已在 to_dataframe() 中格式化为字符串，直接返回
        pass

    return df


def _apply_excel_styles(filepath: str) -> None:
    """
    应用Excel样式

    Args:
        filepath: Excel文件路径
    """
    try:
        wb = load_workbook(filepath)

        # 样式Sheet 1: 渠道维度预算表
        if '渠道维度预算表' in wb.sheetnames:
            ws = wb['渠道维度预算表']
            _style_table1_sheet(ws)

        # 样式Sheet 2: 客群维度汇总表
        if '客群维度汇总表' in wb.sheetnames:
            ws = wb['客群维度汇总表']
            _style_table2_sheet(ws)

        wb.save(filepath)

    except Exception as e:
        # 样式应用失败不应该阻止导出
        print(f"Warning: 应用Excel样式失败: {e}")


def _style_table1_sheet(ws) -> None:
    """
    为Sheet 1应用样式

    Args:
        ws: openpyxl worksheet对象
    """
    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 应用表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 数据行样式
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 汇总行样式 (最后一行)
    total_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    total_font = Font(bold=True, size=11)

    for cell in ws[ws.max_row]:
        cell.fill = total_fill
        cell.font = total_font
        cell.border = thin_border

    # 调整列宽
    for col_idx, col in enumerate(ws.columns, start=1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                pass
        # 设置列宽 (最小12, 最大30)
        adjusted_width = min(max(max_length + 2, 12), 30)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 冻结首行
    ws.freeze_panes = "A2"


def _style_table2_sheet(ws) -> None:
    """
    为Sheet 2应用样式 (保留指标名称中的缩进格式)

    Args:
        ws: openpyxl worksheet对象
    """
    # 表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    # 边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 应用表头样式
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # 数据行样式
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # 第一列样式（已包含缩进）
        if row:
            first_cell = row[0]
            first_cell.alignment = Alignment(horizontal="left", vertical="center")

            # 判断是否为顶层指标（不包含缩进空格）
            if first_cell.value and not str(first_cell.value).startswith(' '):
                first_cell.font = Font(bold=True, size=11)

        # 其他列居中
        for cell in row[1:]:
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 应用边框
        for cell in row:
            cell.border = thin_border

    # 调整列宽
    ws.column_dimensions['A'].width = 35  # 指标列宽一些
    ws.column_dimensions['B'].width = 20  # 交易额列

    # 冻结首行
    ws.freeze_panes = "A2"


# ─────────────────────────────────────────────
# 双引擎对照表
# ─────────────────────────────────────────────

def export_dual_engine(table1, mmm_spends: dict, mmm_loan_amt: float, output):
    """V01 vs MMM 渠道级对比 Excel。

    Parameters
    ----------
    table1 : Table1Result from V01 pipeline
    mmm_spends : dict[channel_name, recommended_spend_万元]
    mmm_loan_amt : float, MMM predicted total loan amount (万元)
    output : file path or BytesIO
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "渠道对比"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["渠道", "V01花费(万元)", "MMM花费(万元)", "花费差异(万元)",
               "V01 T0交易额(千万元)", "V01 CPS"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    channels = [ch for ch in table1.channels if ch.channel_name != "总计"]
    for row_idx, ch in enumerate(channels, 2):
        mmm_spend = mmm_spends.get(ch.channel_name, 0)
        ws.cell(row=row_idx, column=1, value=ch.channel_name)
        ws.cell(row=row_idx, column=2, value=round(ch.expense, 0))
        ws.cell(row=row_idx, column=3, value=round(mmm_spend, 0))
        ws.cell(row=row_idx, column=4, value=round(mmm_spend - ch.expense, 0))
        ws.cell(row=row_idx, column=5, value=round(ch.t0_transaction * 10, 2))
        ws.cell(row=row_idx, column=6, value=f"{(ch.cps_1_8 or 0):.2%}")

    # Summary row
    sum_row = len(channels) + 2
    ws.cell(row=sum_row, column=1, value="合计").font = Font(bold=True)
    ws.cell(row=sum_row, column=2, value=round(table1.total_expense, 0))
    ws.cell(row=sum_row, column=3, value=round(sum(mmm_spends.values()), 0))
    ws.cell(row=sum_row, column=4, value=round(sum(mmm_spends.values()) - table1.total_expense, 0))

    for row in ws.iter_rows(min_row=1, max_row=sum_row, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 18
    ws.freeze_panes = "A2"

    wb.save(output)


# ─────────────────────────────────────────────
# 计算逻辑文档
# ─────────────────────────────────────────────

def export_logic_document(output):
    """导出 V01 计算逻辑手册 (3 sheets)。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def _write_sheet(ws, headers, rows):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for r_idx, row_data in enumerate(rows, 2):
            for c_idx, val in enumerate(row_data, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        for row in ws.iter_rows(min_row=1, max_row=len(rows) + 1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        ws.freeze_panes = "A2"

    # Sheet 1: 计算流程
    ws1 = wb.active
    ws1.title = "计算流程"
    _write_sheet(ws1, ["步骤", "名称", "公式", "说明"], [
        ["①", "预算分配", "渠道花费 = 总预算 × 渠道占比", "按花费结构分配"],
        ["②", "申完量", "申完量 = 花费×10000 ÷ T0申完成本", "花费转化为申请完成量"],
        ["③", "授信量", "授信量 = 申完量 × 1-3过件率", "申请中通过审核的数量"],
        ["④", "T0交易额", "T0 = 花费 ÷ CPS ÷ 10000", "首借当天产生的交易额(亿元)"],
        ["⑤", "M0交易额", "M0 = T0 × M0/T0系数", "首月累计交易额"],
    ])
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 40
    ws1.column_dimensions['D'].width = 30

    # Sheet 2: 参数定义
    ws2 = wb.create_sheet("参数定义")
    _write_sheet(ws2, ["参数名", "单位", "来源", "说明"], [
        ["总预算", "万元", "用户输入", "当月总花费预算"],
        ["渠道花费占比", "%", "用户输入", "各渠道花费占总预算比例"],
        ["1-3 T0过件率", "%", "历史数据/用户输入", "非年龄拒绝口径的1-3天过件率"],
        ["1-8 T0 CPS", "%", "历史数据/用户输入", "首借24h借款金额成本占比"],
        ["T0申完成本", "元", "历史数据", "每笔申请完成的平均成本"],
        ["M0/T0系数", "倍数", "历史计算", "M0与T0交易额的比值(6个月均值)"],
        ["存量M0 CPS", "%", "历史计算", "存量首登客群的CPS(3/6月均值)"],
        ["非初审交易额", "亿元", "用户输入", "非初审客群的预估交易额"],
    ])
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 40

    # Sheet 3: 指标层级
    ws3 = wb.create_sheet("指标层级")
    _write_sheet(ws3, ["层级", "指标", "计算方式"], [
        ["Table1-渠道", "花费(千万元)", "总预算 × 渠道占比 ÷ 10000"],
        ["Table1-渠道", "T0交易额(千万元)", "花费 ÷ CPS ÷ 10000 × 10"],
        ["Table1-渠道", "M0交易额(千万元)", "T0 × M0/T0系数 × 10"],
        ["Table1-渠道", "T0申完量", "花费×10000 ÷ T0申完成本"],
        ["Table1-渠道", "1-3授信量", "申完量 × 1-3过件率"],
        ["Table2-客群", "当月首登M0", "各渠道M0交易额汇总"],
        ["Table2-客群", "存量首登M0", "存量M0费用 ÷ 存量CPS"],
        ["Table2-客群", "总首借交易额", "初审M0 + 存量M0 + T0 + 非初审"],
        ["Table2-客群", "全业务CPS", "总花费 ÷ 总首借交易额"],
    ])
    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 22
    ws3.column_dimensions['C'].width = 40

    wb.save(output)


# ─────────────────────────────────────────────
# MMM 模型报告
# ─────────────────────────────────────────────

def export_mmm_report(model, output):
    """导出 MMM 模型质量报告 Excel。

    Parameters
    ----------
    model : MMMModel from engine/mmm_engine.py
    output : file path or BytesIO
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Sheet 1: 模型质量
    ws1 = wb.active
    ws1.title = "模型质量"
    metrics = [
        ("训练R²", f"{model.r_squared:.4f}"),
        ("测试R²", f"{model.test_r_squared:.4f}" if model.test_r_squared is not None else "N/A"),
        ("NRMSE", f"{model.nrmse:.4f}"),
        ("MAPE (Holdout)", f"{model.mape_holdout:.2%}" if model.mape_holdout is not None else "N/A"),
        ("DW统计量", f"{model.dw_stat:.4f}" if model.dw_stat is not None else "N/A"),
        ("截距", f"{model.intercept:.4f}"),
    ]
    for col_idx, h in enumerate(["指标", "值"], 1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for r_idx, (name, val) in enumerate(metrics, 2):
        ws1.cell(row=r_idx, column=1, value=name)
        ws1.cell(row=r_idx, column=2, value=val)
    for row in ws1.iter_rows(min_row=1, max_row=len(metrics) + 1, max_col=2):
        for cell in row:
            cell.border = thin_border
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 18
    ws1.freeze_panes = "A2"

    # Sheet 2: 渠道参数
    ws2 = wb.create_sheet("渠道参数")
    ch_headers = ["渠道", "Adstock类型", "Theta", "Alpha(Hill)", "Gamma(Hill)", "Beta", "重要度"]
    for col_idx, h in enumerate(ch_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    importance = model.feature_importance or {}
    for r_idx, (ch_name, cp) in enumerate(model.channel_params.items(), 2):
        ws2.cell(row=r_idx, column=1, value=ch_name)
        ws2.cell(row=r_idx, column=2, value=cp.adstock_type)
        ws2.cell(row=r_idx, column=3, value=round(cp.theta, 4))
        ws2.cell(row=r_idx, column=4, value=round(cp.alpha, 4))
        ws2.cell(row=r_idx, column=5, value=round(cp.gamma, 4))
        ws2.cell(row=r_idx, column=6, value=round(cp.beta, 6))
        ws2.cell(row=r_idx, column=7, value=f"{importance.get(ch_name, 0):.2%}")

    n_ch = len(model.channel_params)
    for row in ws2.iter_rows(min_row=1, max_row=n_ch + 1, max_col=len(ch_headers)):
        for cell in row:
            cell.border = thin_border
    for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
        ws2.column_dimensions[col_letter].width = 16
    ws2.freeze_panes = "A2"

    wb.save(output)
