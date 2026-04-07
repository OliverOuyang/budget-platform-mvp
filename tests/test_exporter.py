import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))

import pytest
import tempfile
import shutil
import pandas as pd
from openpyxl import load_workbook

from core.exporter import export_to_excel, _format_dataframe


def test_export_creates_file(sample_table1_result, sample_table2_result):
    """export_to_excel creates a valid xlsx file at the given path."""
    dirpath = tempfile.mkdtemp()
    try:
        output_path = str(Path(dirpath) / "output.xlsx")
        result_path = export_to_excel(
            table1=sample_table1_result,
            table2=sample_table2_result,
            output_path=output_path,
        )
        assert Path(result_path).exists()
        wb = load_workbook(result_path)
        assert '渠道维度预算表' in wb.sheetnames
        assert '客群维度汇总表' in wb.sheetnames
    finally:
        shutil.rmtree(dirpath, ignore_errors=True)


def test_format_dataframe_table1(sample_table1_result):
    """Percentage columns in table1 are formatted as strings with '%'."""
    df = sample_table1_result.to_dataframe()
    formatted = _format_dataframe(df, sheet_type="table1")
    # 花费结构 column should be formatted as percentage string
    if '花费结构' in formatted.columns:
        for val in formatted['花费结构']:
            if val:  # skip empty strings
                assert '%' in str(val)
